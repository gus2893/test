
import json
import boto3
import os
from openpyxl import load_workbook
from io import BytesIO

s3_client = boto3.client('s3')

def lambda_handler(event, context):
    # Get bucket name and file key from environment variables (you can configure these in Lambda console)
    bucket_name = os.getenv('BUCKET_NAME')  # Get bucket name from environment variable
    file_key = os.getenv('FILE_KEY')  # Get file key from environment variable
    output_file_key =  os.getenv('OUTPUT_FILE_KEY')

     # Define reusable Content-Type for .xlsm (macro-enabled Excel files)
    content_type = 'application/vnd.ms-excel.sheet.macroEnabled.12'
    
    # Parse JSON data from request body
    try:
        body = json.loads(event['body'])  # Extract JSON data from the body
        data = body.get('data')  # The actual JSON data to be written
        full_report = body.get('full_report', False)  # Check if the user wants a full report
    except (KeyError, TypeError, json.JSONDecodeError):
        print(f"Error parsing input: {e}")
        return {
            'statusCode': 400,
            'body': json.dumps('Invalid JSON data or missing "data" field.')
        }

    # Download the Excel file from S3 into memory
    try:
        s3_response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
        print(f"Fetching file '{file_key}' from bucket '{bucket_name}'")
        file_content = s3_response['Body'].read()
        wb = load_workbook(filename=BytesIO(file_content), keep_vba=True) # Load Excel file into memory
        ws = wb.active  # Select the active worksheet
    except Exception as e:
        return {
            'statusCode': 500,
            'body': json.dumps(f'Error loading Excel file from S3: {str(e)}')
        }

    # Read the existing headers from the first row of the worksheet
    existing_headers = [cell.value for cell in ws[1] if cell.value]  # Only consider non-empty cells in the first row
    all_keys = set(existing_headers)  # Start with existing headers

        # Collect all keys from the JSON data and add any new ones to the header row
    for entry in data:
        for key in entry.keys():
            if key not in all_keys:
                all_keys.add(key)
                existing_headers.append(key)

    # Now, overwrite the entire header row starting from column A (explicitly starting from column 1)
    for col_idx, header in enumerate(existing_headers, start=1):
        ws.cell(row=1, column=col_idx).value = header

    # Append data dynamically based on the updated headers, ensuring it starts at column A
    for entry in data:
        # Create the row based on the order of the existing headers
        row_data = [entry.get(header, "") for header in existing_headers]
        ws.append(row_data)  # This will append the data starting from column A

    # Check if the full_report flag is set to True
    if full_report:
        # Save the updated workbook to S3
        try:
            output_stream = BytesIO()
            wb.save(output_stream)
            output_stream.seek(0)

            # Upload the updated file back to S3
            s3_client.put_object(Body=output_stream.getvalue(), Bucket=bucket_name, Key=output_file_key, ContentType=content_type)

            return {
                'statusCode': 200,
                'body': json.dumps('Excel file updated and saved to S3 successfully.')
            }

        except Exception as e:
            return {
                'statusCode': 500,
                'body': json.dumps(f'Error saving updated Excel file to S3: {str(e)}')
            }

    else:
        # Return the updated file directly to the user as a binary stream
        try:
            output_stream = BytesIO()
            wb.save(output_stream)
            output_stream.seek(0)

            # Return the binary data directly in the response
            return {
                'statusCode': 200,
                'headers': {
                    'Content-Type': content_type,  # Use the content type variable
                    'Content-Disposition': f'attachment; filename="{output_file_key}"'
                },
                'isBase64Encoded': False,  # Set to False since we're returning raw binary data
                'body': output_stream.getvalue()  # Return the binary content of the Excel file
            }

        except Exception as e:
            return {
                'statusCode': 500,
                'body': json.dumps(f'Error returning updated Excel file: {str(e)}')
            }
