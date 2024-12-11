import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, timedelta
import random

def write_json_to_xlsm(json_data, template_path, output_path):
    """
    Write JSON data to a new worksheet in a macro-enabled Excel (.xlsm) file.
    """
    wb = load_workbook(filename=template_path, keep_vba=True)

    # Select the active worksheet (or a specific sheet by name)
    ws = wb.active

    # Read the existing headers from the first row
    existing_headers = [cell.value for cell in ws[1] if cell.value]  # Only consider non-empty cells in the first row

    # Collect all the keys (fields) from the JSON data
    all_keys = set(existing_headers)  # Start with existing headers

    # Update headers in Excel if there are new keys in the JSON data
    for entry in json_data:
        for key in entry.keys():
            if key not in all_keys:
                all_keys.add(key)
                existing_headers.append(key)
                ws.cell(row=1, column=len(existing_headers)).value = key  # Write new headers starting from column A

    # Append data dynamically based on the updated headers
    for entry in json_data:
        # Create the row based on the order of the existing headers
        row_data = [entry.get(header, "") for header in existing_headers]
        ws.append(row_data)

    # Save the workbook (this will overwrite the existing file)
    wb.save(output_path)
    print(f"Data successfully written to {output_path}")


# Function to generate random integers
def random_int(min_value, max_value):
    return random.randint(min_value, max_value)

# Function to generate random dates
def random_date(start_date, end_date):
    delta = end_date - start_date
    random_days = random.randint(0, delta.days)
    return start_date + timedelta(days=random_days)
def generate_data(entry_count=10):
    data = []

    for _ in range(entry_count):
        entry = {
        }
        data.append(entry)

    return data

# Example usage
if __name__ == "__main__":
    # Sample JSON input
    json_data = generate_data(entry_count=1000)


    # Path to the macro-enabled Excel template
    template_path = "template_with_macros.xlsm"

    # Output path for the new Excel file
    output_path = "output_with_macros.xlsm"

    # Write JSON data to the new worksheet
    write_json_to_xlsm(json_data, template_path, output_path)
